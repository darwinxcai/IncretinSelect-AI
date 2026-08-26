import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from incretinselect.holdout import (
    HoldoutValidationError,
    exact_mean,
    hamming_distance,
    load_design_sequences,
    load_receptor_replicates,
    nearest_training,
    parse_replicate,
)


class HoldoutTests(unittest.TestCase):
    def test_design_sequence_loader_reads_only_p1_to_p15_sheets(self) -> None:
        accessed: list[str] = []

        class FakeSheet:
            def __init__(self, peptide_id: str) -> None:
                self.peptide_id = peptide_id

            def iter_rows(self, **kwargs: object):
                self.assert_cell_range(kwargs)
                sequence = list("A" * 30)
                sequence[0] = "ACDEFGHIKLMNPQRST"[int(self.peptide_id[1:]) - 1]
                yield (self.peptide_id, *sequence)

            @staticmethod
            def assert_cell_range(kwargs: dict[str, object]) -> None:
                if kwargs != {
                    "min_row": 2,
                    "max_row": 2,
                    "min_col": 1,
                    "max_col": 31,
                    "values_only": True,
                }:
                    raise AssertionError(f"Unexpected sequence-cell access: {kwargs}")

        class FakeWorkbook:
            sheetnames = [*list(f"P{index}" for index in range(1, 16)), "5a_GCG", "5b_GLP1"]

            def __getitem__(self, name: str) -> FakeSheet:
                accessed.append(name)
                return FakeSheet(name)

        with patch("openpyxl.load_workbook", return_value=FakeWorkbook()):
            sequences = load_design_sequences("unused.xlsx")
        self.assertEqual(set(sequences), set(f"P{index}" for index in range(1, 16)))
        self.assertEqual(accessed, list(f"P{index}" for index in range(1, 16)))

    def test_censored_replicate_is_not_converted_to_exact_value(self) -> None:
        parsed = parse_replicate(">280000")
        self.assertEqual(parsed, {"status": "right_censored", "threshold_pm": 280000.0})
        self.assertIsNone(exact_mean([parsed]))

    def test_observed_replicates_can_have_exact_mean(self) -> None:
        replicates = [parse_replicate(value) for value in (1.0, 2.0, 3.0)]
        self.assertEqual(exact_mean(replicates), 2.0)

    def test_missing_is_distinct_from_censored(self) -> None:
        self.assertEqual(parse_replicate(None), {"status": "missing"})
        self.assertEqual(parse_replicate(""), {"status": "missing"})

    def test_invalid_replicate_fails(self) -> None:
        with self.assertRaises(HoldoutValidationError):
            parse_replicate("inactive")

    def test_hamming_and_deterministic_nearest_training(self) -> None:
        training = {"b": "AAAT", "a": "AAAC"}
        self.assertEqual(hamming_distance("AAAA", "AAAT"), 1)
        self.assertEqual(nearest_training("AAAA", training), ("a", 1))

    def test_receptor_sheet_identity_is_enforced(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "wrong_receptor.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "data_GCGR"
            worksheet.append(["peptide", "n=1", "n=2", "n=3"])
            workbook.save(path)
            with self.assertRaises(HoldoutValidationError):
                load_receptor_replicates(path, "GLP-1R")


if __name__ == "__main__":
    unittest.main()
