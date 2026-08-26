"""Load the audited same-assay training records used by CPU baselines."""

from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path

from incretinselect.activity import STANDARD_AMINO_ACIDS


class TrainingDataError(ValueError):
    """Raised when training records cannot be joined or interpreted safely."""


@dataclass(frozen=True)
class TrainingRecord:
    peptide_id: str
    raw_sequence: str
    aligned_sequence: str
    gcgr_ec50_pm: float
    glp1r_ec50_pm: float

    @property
    def gcgr_log10_ec50_pm(self) -> float:
        return math.log10(self.gcgr_ec50_pm)

    @property
    def glp1r_log10_ec50_pm(self) -> float:
        return math.log10(self.glp1r_ec50_pm)

    @property
    def selectivity_log10_ratio(self) -> float:
        """log10(GCGR EC50 / GLP-1R EC50); positive favors GLP-1R potency."""

        return self.gcgr_log10_ec50_pm - self.glp1r_log10_ec50_pm


def _header_index(header: tuple, required: set[str], sheet: str) -> dict[str, int]:
    indexes = {str(value): index for index, value in enumerate(header) if value is not None}
    missing = sorted(required - set(indexes))
    if missing:
        raise TrainingDataError(f"{sheet} is missing columns: {', '.join(missing)}")
    return indexes


def _positive_number(value: object, label: str) -> float:
    if isinstance(value, bool):
        raise TrainingDataError(f"{label} must be a positive finite number")
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError) as exc:
        raise TrainingDataError(f"{label} must be a positive finite number") from exc
    if not math.isfinite(number) or number <= 0:
        raise TrainingDataError(f"{label} must be a positive finite number")
    return number


def load_training_records(path: str | Path, expected_records: int | None = None) -> list[TrainingRecord]:
    """Join labels and aligned sequences by peptide ID from one source workbook."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("XLSX parsing requires: pip install -e '.[data]'") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in ("dataset", "alignment"):
        if sheet not in workbook.sheetnames:
            raise TrainingDataError(f"Training workbook is missing {sheet!r}")

    dataset_rows = workbook["dataset"].iter_rows(values_only=True)
    dataset_header = next(dataset_rows, None)
    if dataset_header is None:
        raise TrainingDataError("dataset sheet is empty")
    dataset_indexes = _header_index(
        dataset_header,
        {"pep_ID", "sequence", "EC50_T1", "EC50_T2"},
        "dataset",
    )
    labels: dict[str, tuple[str, float, float]] = {}
    for row in dataset_rows:
        peptide_id = str(row[dataset_indexes["pep_ID"]]).strip()
        if not peptide_id:
            continue
        if peptide_id in labels:
            raise TrainingDataError(f"Duplicate dataset peptide ID: {peptide_id}")
        raw_sequence = str(row[dataset_indexes["sequence"]]).strip().upper()
        invalid = sorted(set(raw_sequence) - STANDARD_AMINO_ACIDS)
        if not raw_sequence or invalid:
            raise TrainingDataError(f"Invalid natural-amino-acid sequence for {peptide_id}")
        labels[peptide_id] = (
            raw_sequence,
            _positive_number(row[dataset_indexes["EC50_T1"]], f"{peptide_id} GCGR EC50"),
            _positive_number(row[dataset_indexes["EC50_T2"]], f"{peptide_id} GLP-1R EC50"),
        )

    alignment_rows = workbook["alignment"].iter_rows(values_only=True)
    alignment_header = next(alignment_rows, None)
    if alignment_header is None:
        raise TrainingDataError("alignment sheet is empty")
    alignment_indexes = _header_index(alignment_header, {"pep_ID", "sequence"}, "alignment")
    alignments: dict[str, str] = {}
    for row in alignment_rows:
        peptide_id = str(row[alignment_indexes["pep_ID"]]).strip()
        if not peptide_id:
            continue
        if peptide_id in alignments:
            raise TrainingDataError(f"Duplicate alignment peptide ID: {peptide_id}")
        sequence = str(row[alignment_indexes["sequence"]]).strip().upper()
        invalid = sorted(set(sequence) - (STANDARD_AMINO_ACIDS | {"-"}))
        if not sequence or invalid:
            raise TrainingDataError(f"Invalid aligned sequence for {peptide_id}")
        alignments[peptide_id] = sequence

    if set(labels) != set(alignments):
        raise TrainingDataError("dataset and alignment peptide ID sets differ")
    lengths = {len(sequence) for sequence in alignments.values()}
    if len(lengths) != 1:
        raise TrainingDataError(f"Aligned sequences do not share one length: {sorted(lengths)}")
    if expected_records is not None and len(labels) != expected_records:
        raise TrainingDataError(f"Expected {expected_records} records; observed {len(labels)}")

    return [
        TrainingRecord(
            peptide_id=peptide_id,
            raw_sequence=labels[peptide_id][0],
            aligned_sequence=alignments[peptide_id],
            gcgr_ec50_pm=labels[peptide_id][1],
            glp1r_ec50_pm=labels[peptide_id][2],
        )
        for peptide_id in sorted(labels)
    ]
