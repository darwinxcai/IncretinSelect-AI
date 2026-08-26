"""Validate the public GCGR/GLP-1R activity workbook without altering it."""

from __future__ import annotations

import argparse
import json
import math
from dataclasses import asdict, dataclass
from numbers import Real
from pathlib import Path
from typing import Any, Iterable

STANDARD_AMINO_ACIDS = frozenset("ACDEFGHIKLMNPQRSTVWY")


@dataclass(frozen=True)
class ValidationIssue:
    severity: str
    code: str
    message: str
    sheet: str | None = None
    row: int | None = None


@dataclass
class ValidationReport:
    workbook: str
    dataset_records: int
    alignment_records: int
    issues: list[ValidationIssue]

    @property
    def passed(self) -> bool:
        return not any(issue.severity == "error" for issue in self.issues)

    def to_dict(self) -> dict[str, Any]:
        return {
            "workbook": self.workbook,
            "passed": self.passed,
            "dataset_records": self.dataset_records,
            "alignment_records": self.alignment_records,
            "error_count": sum(i.severity == "error" for i in self.issues),
            "warning_count": sum(i.severity == "warning" for i in self.issues),
            "issues": [asdict(issue) for issue in self.issues],
        }


def _is_number(value: Any) -> bool:
    return isinstance(value, Real) and not isinstance(value, bool) and math.isfinite(float(value))


def _issue(
    code: str,
    message: str,
    sheet: str,
    row: int | None = None,
    severity: str = "error",
) -> ValidationIssue:
    return ValidationIssue(severity=severity, code=code, message=message, sheet=sheet, row=row)


def _validate_unique_ids(
    rows: list[dict[str, Any]], id_column: str, sheet: str
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    seen: dict[Any, int] = {}
    for position, record in enumerate(rows, start=2):
        value = record.get(id_column)
        if value in (None, ""):
            issues.append(_issue("missing_id", f"Missing {id_column}", sheet, position))
        elif value in seen:
            issues.append(
                _issue(
                    "duplicate_id",
                    f"Duplicate {id_column}={value!r}; first seen on row {seen[value]}",
                    sheet,
                    position,
                )
            )
        else:
            seen[value] = position
    return issues


def validate_rows(
    dataset_rows: Iterable[dict[str, Any]],
    alignment_rows: Iterable[dict[str, Any]],
    config: dict[str, Any],
    workbook: str = "<in-memory>",
    initial_issues: Iterable[ValidationIssue] = (),
) -> ValidationReport:
    """Validate already-loaded workbook records; useful for offline tests."""

    dataset = list(dataset_rows)
    alignment = list(alignment_rows)
    issues = list(initial_issues)
    dataset_sheet = config["dataset_sheet"]
    alignment_sheet = config["alignment_sheet"]
    expected = int(config["expected_records"])

    if len(dataset) != expected:
        issues.append(
            _issue(
                "record_count",
                f"Expected {expected} records, observed {len(dataset)}",
                dataset_sheet,
            )
        )
    if len(alignment) != expected:
        issues.append(
            _issue(
                "record_count",
                f"Expected {expected} records, observed {len(alignment)}",
                alignment_sheet,
            )
        )

    id_column = config["id_column"]
    sequence_column = config["sequence_column"]
    length_column = config["length_column"]
    issues.extend(_validate_unique_ids(dataset, id_column, dataset_sheet))
    issues.extend(_validate_unique_ids(alignment, id_column, alignment_sheet))

    for position, record in enumerate(dataset, start=2):
        sequence = record.get(sequence_column)
        declared_length = record.get(length_column)
        if not isinstance(sequence, str) or not sequence:
            issues.append(_issue("invalid_sequence", "Sequence is empty or non-text", dataset_sheet, position))
        else:
            invalid = sorted(set(sequence.upper()) - STANDARD_AMINO_ACIDS)
            if invalid:
                issues.append(
                    _issue(
                        "invalid_residue",
                        f"Raw sequence contains invalid symbols: {''.join(invalid)}",
                        dataset_sheet,
                        position,
                    )
                )
            if declared_length != len(sequence):
                issues.append(
                    _issue(
                        "length_mismatch",
                        f"Declared length {declared_length!r} != sequence length {len(sequence)}",
                        dataset_sheet,
                        position,
                    )
                )

        for measurement in config["measurements"]:
            ec50_column = measurement["ec50_column"]
            log_column = measurement["log_column"]
            ec50 = record.get(ec50_column)
            log_value = record.get(log_column)
            receptor = measurement["receptor"]
            if not _is_number(ec50) or float(ec50) <= 0:
                issues.append(
                    _issue(
                        "invalid_ec50",
                        f"{receptor} {ec50_column} must be a positive finite number",
                        dataset_sheet,
                        position,
                    )
                )
                continue
            if not _is_number(log_value):
                issues.append(
                    _issue(
                        "invalid_log_ec50",
                        f"{receptor} {log_column} must be a finite number",
                        dataset_sheet,
                        position,
                    )
                )
                continue
            expected_log = math.log10(float(ec50) * 1e-12)
            tolerance = float(config["log_consistency_tolerance"])
            if abs(expected_log - float(log_value)) > tolerance:
                issues.append(
                    _issue(
                        "log_ec50_mismatch",
                        (
                            f"{receptor}: {ec50} pM implies log10(M)={expected_log:.4f}, "
                            f"but {log_column}={log_value}"
                        ),
                        dataset_sheet,
                        position,
                    )
                )

    aligned_length = int(config["aligned_length"])
    for position, record in enumerate(alignment, start=2):
        sequence = record.get(sequence_column)
        declared_length = record.get(length_column)
        if not isinstance(sequence, str) or not sequence:
            issues.append(_issue("invalid_sequence", "Aligned sequence is empty", alignment_sheet, position))
            continue
        invalid = sorted(set(sequence.upper()) - (STANDARD_AMINO_ACIDS | {"-"}))
        if invalid:
            issues.append(
                _issue(
                    "invalid_residue",
                    f"Aligned sequence contains invalid symbols: {''.join(invalid)}",
                    alignment_sheet,
                    position,
                )
            )
        if len(sequence) != aligned_length or declared_length != aligned_length:
            issues.append(
                _issue(
                    "aligned_length_mismatch",
                    (
                        f"Expected aligned length {aligned_length}; observed sequence length "
                        f"{len(sequence)} and declared length {declared_length!r}"
                    ),
                    alignment_sheet,
                    position,
                )
            )

    dataset_ids = {row.get(id_column) for row in dataset if row.get(id_column)}
    alignment_ids = {row.get(id_column) for row in alignment if row.get(id_column)}
    if dataset_ids != alignment_ids:
        issues.append(
            _issue(
                "id_set_mismatch",
                (
                    f"Dataset/alignment ID sets differ: "
                    f"{len(dataset_ids - alignment_ids)} missing from alignment; "
                    f"{len(alignment_ids - dataset_ids)} missing from dataset"
                ),
                alignment_sheet,
            )
        )

    return ValidationReport(
        workbook=workbook,
        dataset_records=len(dataset),
        alignment_records=len(alignment),
        issues=issues,
    )


def _load_sheet(worksheet: Any) -> tuple[list[str], list[dict[str, Any]]]:
    iterator = worksheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        return [], []
    headers = [str(value) if value is not None else "" for value in header_row]
    records: list[dict[str, Any]] = []
    for row in iterator:
        if not any(value is not None for value in row):
            continue
        records.append(
            {header: value for header, value in zip(headers, row, strict=False) if header}
        )
    return [header for header in headers if header], records


def validate_workbook(path: str | Path, config: dict[str, Any]) -> ValidationReport:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment-dependent
        raise RuntimeError("XLSX validation requires: pip install -e '.[data]'") from exc

    workbook_path = Path(path)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    issues: list[ValidationIssue] = []
    loaded: dict[str, tuple[list[str], list[dict[str, Any]]]] = {}
    for sheet_name, required_key in (
        (config["dataset_sheet"], "required_dataset_columns"),
        (config["alignment_sheet"], "required_alignment_columns"),
    ):
        if sheet_name not in workbook.sheetnames:
            issues.append(_issue("missing_sheet", f"Missing sheet {sheet_name!r}", sheet_name))
            loaded[sheet_name] = ([], [])
            continue
        headers, rows = _load_sheet(workbook[sheet_name])
        missing = sorted(set(config[required_key]) - set(headers))
        if missing:
            issues.append(
                _issue(
                    "missing_columns",
                    f"Missing required columns: {', '.join(missing)}",
                    sheet_name,
                )
            )
        loaded[sheet_name] = (headers, rows)

    return validate_rows(
        loaded[config["dataset_sheet"]][1],
        loaded[config["alignment_sheet"]][1],
        config,
        workbook=str(workbook_path),
        initial_issues=issues,
    )


def load_config(path: str | Path) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as handle:
        return json.load(handle)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to training_data.xlsx")
    parser.add_argument("--config", default="configs/activity_schema.json")
    parser.add_argument("--json-output", help="Optional path for a machine-readable report")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    report = validate_workbook(args.workbook, load_config(args.config))
    payload = json.dumps(report.to_dict(), indent=2)
    print(payload)
    if args.json_output:
        output = Path(args.json_output)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(payload + "\n", encoding="utf-8")
    return 0 if report.passed else 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())

