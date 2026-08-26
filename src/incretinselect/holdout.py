"""Freeze and audit the paper's prospective P1--P15 peptide holdout.

The supplement contains right-censored EC50 replicates (for example,
``>280000``). Those values are represented as censoring thresholds, never as
exact observations.
"""

from __future__ import annotations

import math
import statistics
from numbers import Real
from pathlib import Path
from typing import Any, Iterable

from incretinselect.activity import STANDARD_AMINO_ACIDS


class HoldoutValidationError(ValueError):
    """Raised when a prospective-data source fails a required integrity check."""


DESIGN_IDS = tuple(f"P{index}" for index in range(1, 16))
EXPECTED_RECEPTOR_SHEETS = {"GCGR": "data_GCGR", "GLP-1R": "data_GLP-1R"}


def parse_replicate(value: Any) -> dict[str, Any]:
    """Parse one EC50 replicate without erasing censoring or missingness."""

    if value is None or (isinstance(value, str) and not value.strip()):
        return {"status": "missing"}
    if isinstance(value, Real) and not isinstance(value, bool):
        number = float(value)
        if not math.isfinite(number) or number <= 0:
            raise HoldoutValidationError(f"EC50 replicate must be positive and finite: {value!r}")
        return {"status": "observed", "value_pm": number}
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if text.startswith(">"):
            try:
                threshold = float(text[1:].strip())
            except ValueError as exc:
                raise HoldoutValidationError(f"Invalid censoring threshold: {value!r}") from exc
            if not math.isfinite(threshold) or threshold <= 0:
                raise HoldoutValidationError(f"Invalid censoring threshold: {value!r}")
            return {"status": "right_censored", "threshold_pm": threshold}
    raise HoldoutValidationError(f"Unsupported EC50 replicate value: {value!r}")


def exact_mean(replicates: Iterable[dict[str, Any]]) -> float | None:
    values = list(replicates)
    if not values or any(item["status"] != "observed" for item in values):
        return None
    return statistics.fmean(float(item["value_pm"]) for item in values)


def load_design_sequences(path: str | Path) -> dict[str, str]:
    """Read P1--P15 aligned sequences from the Figure 5 source-data workbook."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("XLSX parsing requires: pip install -e '.[data]'") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    missing = sorted(set(DESIGN_IDS) - set(workbook.sheetnames))
    if missing:
        raise HoldoutValidationError(f"Missing design sheets: {', '.join(missing)}")

    sequences: dict[str, str] = {}
    for design_id in DESIGN_IDS:
        row = next(
            workbook[design_id].iter_rows(
                min_row=2,
                max_row=2,
                min_col=1,
                max_col=31,
                values_only=True,
            ),
            None,
        )
        if row is None or str(row[0]).strip() != design_id:
            raise HoldoutValidationError(f"Sheet {design_id} row 2 does not identify {design_id}")
        sequence = "".join(str(value).strip().upper() for value in row[1:31] if value not in (None, ""))
        invalid = sorted(set(sequence) - (STANDARD_AMINO_ACIDS | {"-"}))
        if len(sequence) != 30 or invalid:
            raise HoldoutValidationError(
                f"{design_id} must be a 30-position aligned natural-amino-acid sequence"
            )
        sequences[design_id] = sequence
    if len(set(sequences.values())) != len(sequences):
        raise HoldoutValidationError("Prospective design sequences are not unique")
    return sequences


def load_receptor_replicates(path: str | Path, receptor: str) -> dict[str, list[dict[str, Any]]]:
    """Load three prospective EC50 replicates for one explicitly named receptor sheet."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("XLSX parsing requires: pip install -e '.[data]'") from exc

    expected_sheet = EXPECTED_RECEPTOR_SHEETS.get(receptor)
    if expected_sheet is None:
        raise HoldoutValidationError(f"Unsupported receptor: {receptor!r}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if expected_sheet not in workbook.sheetnames:
        raise HoldoutValidationError(
            f"Expected sheet {expected_sheet!r} for {receptor}; observed {workbook.sheetnames!r}"
        )
    worksheet = workbook[expected_sheet]
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header is None or tuple(header[:4]) != ("peptide", "n=1", "n=2", "n=3"):
        raise HoldoutValidationError(f"Unexpected replicate columns in {expected_sheet}")

    records: dict[str, list[dict[str, Any]]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        peptide = str(row[0]).strip() if row and row[0] is not None else ""
        if peptide not in DESIGN_IDS:
            continue
        if peptide in records:
            raise HoldoutValidationError(f"Duplicate {peptide} row in {expected_sheet}")
        records[peptide] = [parse_replicate(value) for value in row[1:4]]
    missing = sorted(set(DESIGN_IDS) - set(records))
    if missing:
        raise HoldoutValidationError(f"Missing {receptor} rows: {', '.join(missing)}")
    return records


def load_training_alignment(path: str | Path) -> dict[str, str]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("XLSX parsing requires: pip install -e '.[data]'") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "alignment" not in workbook.sheetnames:
        raise HoldoutValidationError("Training workbook is missing the alignment sheet")
    rows = workbook["alignment"].iter_rows(values_only=True)
    header = list(next(rows, ()))
    try:
        id_index = header.index("pep_ID")
        sequence_index = header.index("sequence")
    except ValueError as exc:
        raise HoldoutValidationError("Alignment sheet lacks pep_ID or sequence") from exc
    result: dict[str, str] = {}
    for row in rows:
        peptide_id = str(row[id_index]).strip() if row[id_index] is not None else ""
        sequence = str(row[sequence_index]).strip().upper() if row[sequence_index] else ""
        if peptide_id and sequence:
            if peptide_id in result:
                raise HoldoutValidationError(f"Duplicate training peptide ID: {peptide_id}")
            if len(sequence) != 30:
                raise HoldoutValidationError(f"Training alignment {peptide_id} is not length 30")
            result[peptide_id] = sequence
    if len(result) != 125:
        raise HoldoutValidationError(f"Expected 125 training alignments; observed {len(result)}")
    return result


def hamming_distance(first: str, second: str) -> int:
    if len(first) != len(second):
        raise HoldoutValidationError("Hamming distance requires equal-length aligned sequences")
    return sum(left != right for left, right in zip(first, second, strict=True))


def nearest_training(sequence: str, training: dict[str, str]) -> tuple[str, int]:
    if not training:
        raise HoldoutValidationError("Training alignment set is empty")
    distance, peptide_id = min(
        (hamming_distance(sequence, train_sequence), train_id)
        for train_id, train_sequence in training.items()
    )
    return peptide_id, distance


def design_group(design_id: str) -> str:
    index = int(design_id[1:])
    if index <= 5:
        return "dual"
    if index <= 10:
        return "gcgr_selective"
    return "glp1r_selective"


def build_holdout(
    sequences: dict[str, str],
    gcgr: dict[str, list[dict[str, Any]]],
    glp1r: dict[str, list[dict[str, Any]]],
    training: dict[str, str],
) -> dict[str, Any]:
    expected = set(DESIGN_IDS)
    if set(sequences) != expected or set(gcgr) != expected or set(glp1r) != expected:
        raise HoldoutValidationError("Sequences and both receptor tables must each contain P1--P15")

    training_sequences = set(training.values())
    records = []
    for design_id in DESIGN_IDS:
        nearest_id, nearest_distance = nearest_training(sequences[design_id], training)
        records.append(
            {
                "peptide_id": design_id,
                "design_group": design_group(design_id),
                "aligned_sequence": sequences[design_id],
                "aligned_length": 30,
                "nearest_training_peptide_id": nearest_id,
                "nearest_training_hamming_distance": nearest_distance,
                "gcgr_ec50_replicates": gcgr[design_id],
                "gcgr_exact_mean_pm": exact_mean(gcgr[design_id]),
                "glp1r_ec50_replicates": glp1r[design_id],
                "glp1r_exact_mean_pm": exact_mean(glp1r[design_id]),
            }
        )

    return {
        "schema_version": 1,
        "endpoint": "cAMP accumulation EC50",
        "unit": "pM",
        "endpoint_warning": "Functional potency, not binding affinity; censored values are not exact.",
        "training_records": len(training),
        "holdout_records": len(records),
        "exact_sequence_overlaps": sum(
            record["aligned_sequence"] in training_sequences for record in records
        ),
        "records": records,
    }
